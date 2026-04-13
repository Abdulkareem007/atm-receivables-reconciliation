"""
ATM Receivables GL Reconciliation — Python Script
==================================================
Automates the reconciliation of 5 ATM receivable GL accounts
for a Nigerian bank branch using pandas and openpyxl.

Inputs:
    - Master Excel workbook (one sheet per GL account, proofed to prior period)
    - GL Activity Report files (one .xlsx per GL account, exported from core banking)

Process:
    - Extracts RRN from transaction descriptions using regex
    - Appends new transactions to each GL sheet
    - Recalculates proof balance vs system balance
    - Flags any non-zero difference

Output:
    - Updated master workbook with new PROOF/SYSTEM/DIFFERENCE footer rows
    - Console summary showing balance status per GL account

Usage:
    python reconcile.py

Author: Abdulkareem Muazu
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font


# ── GL account → sheet name mapping ──────────────────────────────
GL_SHEET_MAP = {
    '119110010': '16436-119110010',   # ISW ATM Settlement
    '119110038': '16484-119110038',   # MC Domestic ATM
    '119130021': '16533-119130021',   # Appzone ZS ATM
    '119110026': '16459-119110026',   # VISA / V-Pay
    '119110093': '119110093',         # AFRIGO
}

GL_FILES = {
    '119110010': 'data/Gl_Activity_Report__40_.xlsx',
    '119110038': 'data/Gl_Activity_Report__41_.xlsx',
    '119130021': 'data/Gl_Activity_Report__42_.xlsx',
    '119110026': 'data/Gl_Activity_Report__43_.xlsx',
    '119110093': 'data/Gl_Activity_Report__44_.xlsx',
}

MASTER_FILE = 'data/ATM_AS_AT_31ST_MARCH_2026.xlsx'
OUTPUT_FILE = 'output/ATM_RECONCILED.xlsx'
PROOF_DATE  = '12/04/2026'


def extract_rrn(desc):
    """Extract RRN integer from a GL transaction description string.

    Handles all description formats produced by the core banking system:
    - Old ATI:    ISW|GL|RRN|...  /  MC|GL|RRN|...
    - New ATI:    RRN|GL|TERMINAL|...
    - ZS ATM:     ZS ATM-RRN-...
    - EDB settle: RRN-AFRIGO ATM SETT-...
    - ATM WDL:    ATM WDL-RRN-...
    - Reversals:  ***RSVL / RVSL prefix stripped before matching
    """
    if not desc or not isinstance(desc, str):
        return None
    d = re.sub(r'^\*+', '', desc.strip())
    d = re.sub(r'^(?:RSVL|RVSL)\s+', '', d, flags=re.IGNORECASE)

    patterns = [
        r'^(?:ISW|MC|VC|VGATE)\|\d+\|0*(\d{6,})\|',   # Old ATI
        r'^0*(\d{6,})\|',                               # New ATI
        r'^ZS ATM[-\s]0*(\d{6,})',                      # ZS ATM
        r'^ATM WDL-0*(\d+)',                             # ATM WDL
        r'^0*(\d{6,})-',                                 # EDB settlement
    ]
    for pattern in patterns:
        m = re.match(pattern, d, re.IGNORECASE)
        if m:
            return int(m.group(1))
    return None


def load_gl_report(filepath):
    """Load a GL Activity Report, skipping the header block and opening balance row."""
    df = pd.read_excel(filepath, header=17, skiprows=range(18, 20))
    df = df.dropna(how='all')
    df = df[df['TRN CODE'].notna() & (df['TRN CODE'].astype(str).str.strip() != '')]
    df['RRN'] = df['DESCRIPTION'].apply(extract_rrn)
    return df


def get_system_balance(df):
    """Get the closing system balance from the last row of the BALANCE column."""
    last = df['BALANCE'].dropna().iloc[-1]
    return float(str(last).replace(',', '').replace(' ', ''))


def remove_footer_rows(ws):
    """Remove existing PROOF/SYSTEM/DIFFERENCE footer rows from a sheet."""
    keywords = ['PROOF BALANCE', 'SYSTEM BALANCE', 'DIFFERENCE']
    footer_rows = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and any(k in cell.value.upper() for k in keywords):
                footer_rows.append(row[0].row)
                break
    for fr in sorted(footer_rows, reverse=True):
        ws.delete_rows(fr)


def append_transactions(ws, df, next_row):
    """Append GL transactions to the target sheet, skipping duplicates by RRN."""
    existing_rrns = set()
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0]:
            existing_rrns.add(str(row[0]))

    added = 0
    for _, t in df.iterrows():
        rrn = t.get('RRN')
        if rrn and str(rrn) in existing_rrns:
            continue
        ws.cell(next_row, 1, rrn)
        ws.cell(next_row, 2, str(t.get('CREATE DATE', '')).strip())
        ws.cell(next_row, 3, str(t.get('EFFECTIVE DATE', '')).strip())
        ws.cell(next_row, 4, str(t.get('TRN CODE', '')).strip())
        ws.cell(next_row, 5, str(t.get('DESCRIPTION', '')).strip())
        ws.cell(next_row, 6, float(t.get('AMOUNT', 0) or 0))
        ws.cell(next_row, 7, float(t.get('DEBIT', 0) or 0))
        ws.cell(next_row, 8, float(t.get('CREDIT', 0) or 0))
        ws.cell(next_row, 9, f'=H{next_row}-G{next_row}')
        ws.cell(next_row, 10, str(t.get('POSTER', '')).strip())
        ws.cell(next_row, 11, str(t.get('BRANCH', '')).strip())
        next_row += 1
        added += 1
    return next_row - 1, added


def add_footer(ws, last_data_row, sys_bal, proof_date):
    """Write PROOF / SYSTEM / DIFFERENCE footer rows."""
    pr, sr, dr = last_data_row + 1, last_data_row + 2, last_data_row + 3
    ws.cell(pr, 5, f'PROOF BALANCE AS AT {proof_date}')
    ws.cell(pr, 9, f'=SUM(I2:I{last_data_row})')
    ws.cell(sr, 5, f'SYSTEM BALANCE AS AT {proof_date}')
    ws.cell(sr, 9, sys_bal)
    ws.cell(dr, 5, 'DIFFERENCE')
    ws.cell(dr, 9, f'=I{pr}-I{sr}')
    for r, c in [(pr,5),(pr,9),(sr,5),(sr,9),(dr,5),(dr,9)]:
        ws.cell(r, c).font = Font(bold=True)


def main():
    wb = load_workbook(MASTER_FILE)
    print(f'\n{"="*55}')
    print(f'  ATM RECEIVABLES RECONCILIATION  |  As at {PROOF_DATE}')
    print(f'{"="*55}')

    results = []
    for gl_no, sheet_name in GL_SHEET_MAP.items():
        ws  = wb[sheet_name]
        df  = load_gl_report(GL_FILES[gl_no])
        sys_bal = get_system_balance(df)

        remove_footer_rows(ws)
        last_row = max((r[0].row for r in ws.iter_rows() if any(c.value for c in r)), default=1)
        last_data_row, added = append_transactions(ws, df, last_row + 1)
        add_footer(ws, last_data_row, sys_bal, PROOF_DATE)

        results.append({
            'GL': gl_no, 'Sheet': sheet_name,
            'Added': added, 'SysBal': sys_bal
        })

    wb.save(OUTPUT_FILE)

    print(f'\n  {"GL Account":<12} {"Sheet":<22} {"Rows Added":>10} {"System Bal":>16}')
    print(f'  {"-"*62}')
    for r in results:
        print(f'  {r["GL"]:<12} {r["Sheet"]:<22} {r["Added"]:>10} {r["SysBal"]:>16,.2f}')
    print(f'\n  Output saved → {OUTPUT_FILE}')
    print(f'{"="*55}\n')


if __name__ == '__main__':
    main()
